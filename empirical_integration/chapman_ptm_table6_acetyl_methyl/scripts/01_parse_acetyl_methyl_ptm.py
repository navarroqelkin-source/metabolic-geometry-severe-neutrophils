"""01 — Parse Chapman Table_6 sheets Acetylation and Methylation into structured PTM entries.

Real columns only (Protein Accession, Protein name, Peptide, PTM). Records the modified residue
site (mass shift) and the PTM type. Rows with empty peptide AND empty PTM are skipped (no fabrication).
Writes CHAPMAN_TABLE6_ACETYL_METHYL_INDEX.tsv and CHAPMAN_TABLE6_PARSED_PTM.tsv.
"""
import csv
import os
import re

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EMP = os.path.dirname(ROOT)
TABLE6 = os.path.join(EMP, "chapman_supplement_tier2_extraction", "raw", "Table_6.XLSX")
SHEETS = {"Acetylation": ("Acetylation", "+42.01"), "Methylation": ("Methylation", "+14.02/+28.03")}
ACC_RE = re.compile(r"^([A-Z0-9]+)\|([A-Za-z0-9_]+)(?:\s*\(([^)]+)\))?")
ALLOWED = "structured acetylated/methylated NET/release material composition (Tier 2 PTM)"
PROHIBITED = ("enzyme activity (HAT/HDAC/PRMT); acetylation/methylation rate; NETosis rate; "
              "NET clearance; pathogenicity; causal mechanism; sample-level fusion")


def parse_acc(s):
    m = ACC_RE.match((s or "").strip())
    if not m:
        return "", "", (s or "").strip()
    uni, entry, paren = m.group(1), m.group(2), m.group(3)
    return uni, entry, (paren or entry or "").strip()


def site_for(peptide, ptm_type):
    # residue immediately preceding the relevant mass token
    if ptm_type == "Acetylation":
        mm = re.search(r"([A-Z])\(\+42\.01", peptide or "")
        return f"{mm.group(1)}(+42.01_acetyl)" if mm else ("N-term(+42.01_acetyl)" if "+42.01" in (peptide or "") else "")
    mm = re.search(r"([A-Z])\(\+(?:14\.02|28\.03)", peptide or "")
    return f"{mm.group(1)}(methyl)" if mm else ("methyl" if ("+14.02" in (peptide or "") or "+28.03" in (peptide or "")) else "")


def find_header(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if any(v.lower() == "peptide" for v in vals):
            return i, vals
    return None, None


def main():
    if not os.path.exists(TABLE6):
        print("[01] Table_6 missing")
        return
    wb = openpyxl.load_workbook(TABLE6, read_only=True, data_only=True)
    idx_rows, out = [], []
    for sheet, (ptm_type, shift) in SHEETS.items():
        if sheet not in wb.sheetnames:
            idx_rows.append({"sheet_name": sheet, "ptm_type": ptm_type, "n_data_rows": "0",
                             "mass_shift": shift, "parse_status": "SHEET_MISSING"})
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        hidx, header = find_header(ws)
        col = {h.lower(): j for j, h in enumerate(header)} if header else {}
        j_acc = next((col[k] for k in col if "accession" in k), 0)
        j_name = next((col[k] for k in col if k == "protein name" or "name" in k), 1)
        j_pep = next((col[k] for k in col if "peptide" in k), 2)
        j_ptm = next((col[k] for k in col if k == "ptm" or "modification" in k), 3)
        n_data = 0
        for r in rows[hidx + 1:]:
            if not r or all(c is None for c in r):
                continue
            pep = str(r[j_pep]).strip() if j_pep < len(r) and r[j_pep] is not None else ""
            ptm = str(r[j_ptm]).strip() if j_ptm < len(r) and r[j_ptm] is not None else ""
            if not pep and not ptm:
                continue  # skip malformed/continuation rows
            if pep.lower() == "peptide" or ptm.lower() == "ptm":
                continue  # skip repeated header rows within the sheet
            acc_raw = str(r[j_acc]).strip() if j_acc < len(r) and r[j_acc] is not None else ""
            if acc_raw.lower() in ("protein accession", "accession"):
                continue
            name = str(r[j_name]).strip() if j_name < len(r) and r[j_name] is not None else ""
            uni, entry, gene = parse_acc(acc_raw)
            n_data += 1
            out.append({
                "source_id": "Chapman_PXD011796", "ptm_type": ptm_type,
                "protein_or_peptide": gene or entry or (pep[:24] if pep else uni),
                "gene_symbol": gene, "uniprot_accession": uni,
                "modification_or_ptm": ptm or ptm_type,
                "site_or_position": site_for(pep, ptm_type),
                "condition_or_group": "RA_SLE (>=22/23 samples)",
                "reported_metric_or_status": (f"protein={name}" if name else "") + f"; mass_shift={shift}",
                "table_source": f"Table_6.XLSX#{sheet}",
                "evidence_tier": "Tier_2",
                "allowed_claim": ALLOWED, "prohibited_claim": PROHIBITED,
            })
        idx_rows.append({"sheet_name": sheet, "ptm_type": ptm_type, "n_data_rows": str(n_data),
                         "mass_shift": shift, "parse_status": "PARSED"})

    with open(os.path.join(ROOT, "CHAPMAN_TABLE6_ACETYL_METHYL_INDEX.tsv"), "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=["sheet_name", "ptm_type", "n_data_rows", "mass_shift", "parse_status"],
                           delimiter="\t", lineterminator="\n")
        w.writeheader(); w.writerows(idx_rows)
    header = ["source_id", "ptm_type", "protein_or_peptide", "gene_symbol", "uniprot_accession",
              "modification_or_ptm", "site_or_position", "condition_or_group",
              "reported_metric_or_status", "table_source", "evidence_tier", "allowed_claim", "prohibited_claim"]
    with open(os.path.join(ROOT, "CHAPMAN_TABLE6_PARSED_PTM.tsv"), "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=header, delimiter="\t", extrasaction="ignore", lineterminator="\n")
        w.writeheader(); w.writerows(out)
    by = {}
    for r in out:
        by[r["ptm_type"]] = by.get(r["ptm_type"], 0) + 1
    print(f"[01] parsed PTM entries={len(out)} by_type={by}")
    for t in ("Acetylation", "Methylation"):
        prots = sorted({r["gene_symbol"] for r in out if r["ptm_type"] == t and r["gene_symbol"]})
        print(f"     {t}: {by.get(t,0)} entries / {len(prots)} proteins e.g. {prots[:10]}")


if __name__ == "__main__":
    main()
