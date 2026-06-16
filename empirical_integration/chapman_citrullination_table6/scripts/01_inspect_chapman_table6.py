"""01 — Inspect Chapman Table_6: enumerate sheets, real dimensions and column names.

No structural assumptions — reads the actual header row of each sheet. Writes CHAPMAN_TABLE6_INDEX.tsv.
"""
import os

import openpyxl

import _cit_common as C


def header_row(ws):
    """Return (header_index, header_list): the first row that looks like a header (has 'Peptide' or 'Accession')."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True)):
        vals = [str(c).strip() if c is not None else "" for c in row]
        low = [v.lower() for v in vals]
        # real header row: a cell is exactly 'peptide' or 'protein accession' (not the title paragraph)
        if any(v in ("peptide", "protein accession", "accession") for v in low):
            return i, [v for v in vals if v]
    return 0, []


def main():
    if not os.path.exists(C.TABLE6):
        C.write_tsv(os.path.join(C.ROOT, "CHAPMAN_TABLE6_INDEX.tsv"),
                    ["source_id", "file_name", "sheet_name", "n_rows", "n_columns",
                     "column_names", "table_interpretation", "parse_status", "required_action"],
                    [{"source_id": "Chapman_PXD011796", "file_name": "Table_6.XLSX",
                      "sheet_name": "(missing)", "parse_status": "FILE_MISSING",
                      "required_action": "re-run chapman_supplement_tier2_extraction download"}])
        print("[01] Table_6.XLSX missing")
        return
    wb = openpyxl.load_workbook(C.TABLE6, read_only=True, data_only=True)
    rows = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        hidx, cols = header_row(ws)
        interp = ("citrullinated peptides per protein with NET-induction stimulus counts"
                  if sn.lower() == "citrullination" else f"{sn} PTM (out of scope)")
        rows.append({
            "source_id": "Chapman_PXD011796", "file_name": "Table_6.XLSX", "sheet_name": sn,
            "n_rows": ws.max_row, "n_columns": ws.max_column,
            "column_names": "; ".join(cols),
            "table_interpretation": interp,
            "parse_status": "PARSE_TARGET" if sn.lower() == "citrullination" else "CONTEXT_ONLY",
            "required_action": "parse citrullinated entries" if sn.lower() == "citrullination" else "none",
        })
    C.write_tsv(os.path.join(C.ROOT, "CHAPMAN_TABLE6_INDEX.tsv"),
                ["source_id", "file_name", "sheet_name", "n_rows", "n_columns",
                 "column_names", "table_interpretation", "parse_status", "required_action"], rows)
    print(f"[01] sheets={wb.sheetnames}")
    for r in rows:
        print(f"     {r['sheet_name']}: {r['n_rows']}x{r['n_columns']} cols=[{r['column_names']}] -> {r['parse_status']}")


if __name__ == "__main__":
    main()
