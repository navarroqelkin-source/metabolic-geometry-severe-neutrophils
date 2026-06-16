"""02 — Parse the Citrullination sheet of Chapman Table_6.

Extracts only real columns present (Protein Accession, Protein name, Peptide, A23187 count,
PMA count). Records the citrullinated residue site (+0.98) and per-stimulus detection counts.
No fabricated entries. Writes CHAPMAN_TABLE6_PARSED_CITRULLINATION.tsv (or NOT_STRUCTURABLE).
"""
import os

import openpyxl

import _cit_common as C


def find_header(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if any(v.lower() == "peptide" for v in vals):
            return i, vals
    return None, None


def main():
    if not os.path.exists(C.TABLE6):
        print("[02] Table_6 missing — NOT_STRUCTURABLE")
        C.write_tsv(os.path.join(C.ROOT, "CHAPMAN_TABLE6_PARSED_CITRULLINATION.tsv"),
                    ["source_id"], [])
        return
    wb = openpyxl.load_workbook(C.TABLE6, read_only=True, data_only=True)
    if C.CIT_SHEET not in wb.sheetnames:
        print("[02] no Citrullination sheet — NOT_STRUCTURABLE")
        C.write_tsv(os.path.join(C.ROOT, "CHAPMAN_TABLE6_PARSED_CITRULLINATION.tsv"), ["source_id"], [])
        return
    ws = wb[C.CIT_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    hidx, header = find_header(ws)
    if hidx is None:
        print("[02] no header found — NOT_STRUCTURABLE")
        C.write_tsv(os.path.join(C.ROOT, "CHAPMAN_TABLE6_PARSED_CITRULLINATION.tsv"), ["source_id"], [])
        return
    col = {h.lower(): j for j, h in enumerate(header)}
    j_acc = next((col[k] for k in col if "accession" in k), 0)
    j_name = next((col[k] for k in col if k.startswith("protein name") or k == "protein name"), 1)
    j_pep = next((col[k] for k in col if "peptide" in k), 2)
    j_a23 = next((col[k] for k in col if "a23187" in k), None)
    j_pma = next((col[k] for k in col if "pma" in k), None)

    out = []
    for r in rows[hidx + 1:]:
        if not r or all(c is None for c in r):
            continue
        acc_raw = str(r[j_acc]).strip() if j_acc < len(r) and r[j_acc] is not None else ""
        pep = str(r[j_pep]).strip() if j_pep < len(r) and r[j_pep] is not None else ""
        if not acc_raw and not pep:
            continue
        uni, entry, gene = C.parse_accession(acc_raw)
        name = str(r[j_name]).strip() if j_name < len(r) and r[j_name] is not None else ""
        a23 = str(r[j_a23]).strip() if j_a23 is not None and j_a23 < len(r) and r[j_a23] is not None else ""
        pma = str(r[j_pma]).strip() if j_pma is not None and j_pma < len(r) and r[j_pma] is not None else ""
        site = C.citrulline_site(pep)
        metric = f"A23187_count={a23 or 'NA'}; PMA_count={pma or 'NA'}"
        out.append({
            "source_id": "Chapman_PXD011796",
            "protein_or_peptide": gene or entry or (pep[:24] if pep else uni),
            "gene_symbol": gene, "uniprot_accession": uni,
            "modification_or_ptm": "Citrullination (Arg +0.98)",
            "site_or_position": site,
            "condition_or_group": "NET_induction(A23187/PMA)",
            "reported_metric_or_status": metric + (f"; protein={name}" if name else ""),
            "table_source": "Table_6.XLSX#Citrullination",
            "evidence_tier": "Tier_2",
            "allowed_claim": C.ALLOWED, "prohibited_claim": C.PROHIBITED,
        })
    if not out:
        print("[02] no citrullinated rows parsed — NOT_STRUCTURABLE")
        C.write_tsv(os.path.join(C.ROOT, "CHAPMAN_TABLE6_PARSED_CITRULLINATION.tsv"), ["source_id"], [])
        return
    C.write_tsv(os.path.join(C.ROOT, "CHAPMAN_TABLE6_PARSED_CITRULLINATION.tsv"),
                ["source_id", "protein_or_peptide", "gene_symbol", "uniprot_accession",
                 "modification_or_ptm", "site_or_position", "condition_or_group",
                 "reported_metric_or_status", "table_source", "evidence_tier",
                 "allowed_claim", "prohibited_claim"], out)
    n_genes = len({r["gene_symbol"] for r in out if r["gene_symbol"]})
    print(f"[02] parsed citrullinated entries={len(out)} unique_proteins={n_genes}")
    sample = sorted({r["gene_symbol"] for r in out if r["gene_symbol"]})[:10]
    print(f"     proteins e.g.: {sample}")


if __name__ == "__main__":
    main()
